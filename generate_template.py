from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from database import SessionLocal
import models
import os


def generate_template(subtypes: dict):
    db = SessionLocal()
    try:
        points = db.query(models.Point).order_by(models.Point.owner, models.Point.name).all()
    finally:
        db.close()

    # Разделяем точки по владельцам
    afonin_points = [p for p in points if p.owner == "ИП Афонин"]
    soldatov_points = [p for p in points if p.owner == "ИП Солдатов"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон"

    # ================= Стили =================
    thick_side = Side(border_style="thick", color="000000")
    thin_side = Side(border_style="thin", color="000000")
    thick_border = Border(top=thick_side, left=thick_side, right=thick_side, bottom=thick_side)
    thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    afonin_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    # ================= Колонки =================
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 6
    col_idx = 3
    for p in afonin_points + soldatov_points:
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
        col_idx += 1

    # ================= Высота строк =================
    ws.row_dimensions[1].height = 180
    for i in range(2, 200):
        ws.row_dimensions[i].height = 25

    # ================= Заголовки точек =================
    col_idx = 3
    for p in afonin_points:
        cell = ws.cell(row=1, column=col_idx, value=p.name)
        cell.fill = afonin_fill
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(textRotation=90, vertical="bottom", horizontal="center")
        cell.border = thick_border
        col_idx += 1

    for p in soldatov_points:
        cell = ws.cell(row=1, column=col_idx, value=p.name)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(textRotation=90, vertical="bottom", horizontal="center")
        cell.border = thick_border
        col_idx += 1

    # ================= Товары =================
    current_row = 2
    max_col = 2 + len(points)

    def add_subtype(subtype_name, unit, items, skip_header=False):
        nonlocal current_row

        if not skip_header:
            # Жёлтая строка подтипа
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
            cell = ws.cell(row=current_row, column=1, value=subtype_name)
            cell.fill = yellow_fill
            cell.font = Font(bold=True, size=13)
            cell.alignment = Alignment(horizontal="left", vertical="center")

            for col in range(1, max_col + 1):
                ws.cell(row=current_row, column=col).border = thick_border

            current_row += 1

        # Товары
        for item in items:
            ws.cell(row=current_row, column=1, value=item).font = Font(size=12)
            ws.cell(row=current_row, column=2, value=unit).font = Font(size=12)
            for col in range(1, max_col + 1):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

    # ================= Обработка подтипов =================
    for subtype_name, data in subtypes.items():
        unit = data.get("unit", "")
        items = data.get("items", [])
        skip_header = subtype_name == "unsub"
        add_subtype(subtype_name, unit, items, skip_header=skip_header)

    # ================= Сохранение =================
    os.makedirs("exports", exist_ok=True)
    filepath = os.path.join("exports", "template.xlsx")
    wb.save(filepath)
    print(f"Шаблон сохранен: {filepath}")


def generate_template_horizontal(subtypes: dict):
    """Горизонтальный шаблон (точки в строках, товары сверху).
    Работает только если subtypes содержит один ключ — 'unsub'.
    """

    if list(subtypes.keys()) != ["unsub"]:
        print("❌ Горизонтальный шаблон можно создать только для ключа 'unsub'")
        return

    db = SessionLocal()
    try:
        points = db.query(models.Point).order_by(models.Point.owner, models.Point.name).all()
    finally:
        db.close()

    afonin_points = [p for p in points if p.owner == "ИП Афонин"]
    soldatov_points = [p for p in points if p.owner == "ИП Солдатов"]

    products = subtypes["unsub"]["items"]
    unit = subtypes["unsub"].get("unit", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Горизонтальный шаблон"

    # ===== Стили =====
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    afonin_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    # ===== Колонки =====
    ws.column_dimensions["A"].width = 34  # ширина колонки с точками
    for i in range(len(products)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 11  # остальные

    # ===== Высота строк =====
    ws.row_dimensions[1].height = 45.75  # строка с товарами
    for i in range(2, len(points) + 3):
        ws.row_dimensions[i].height = 15

    # ===== Заголовки товаров =====
    for col, product in enumerate(products, start=2):
        cell = ws.cell(row=1, column=col, value=product)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # ===== Точки =====
    row = 2
    for p in afonin_points + soldatov_points:
        cell = ws.cell(row=row, column=1, value=p.name)
        cell.font = Font(bold=True, size=12)
        if p.owner == "ИП Афонин":
            cell.fill = afonin_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        row += 1

    # ===== Границы таблицы =====
    for r in range(1, row):
        for c in range(1, len(products) + 2):
            ws.cell(row=r, column=c).border = thin_border

    # ===== Сохранение =====
    os.makedirs("exports", exist_ok=True)
    filepath = os.path.join("exports", "template_horizontal.xlsx")
    wb.save(filepath)
    print(f"Горизонтальный шаблон сохранен: {filepath}")


if __name__ == "__main__":
    subtypes = {
        "unsub": {
            "unit": "шт.",
            "items": [
                "масло растительное",
                "соль",
                "яйцо",
                "чеснок",
                "сок яблоко",
                "сок персик/абрик",
                "сок мульти",
                "сок вишня",
                "сок апельсин",
                "чай черн",
                "чай зелен"
            ]
        }
    }

    generate_template(subtypes)
