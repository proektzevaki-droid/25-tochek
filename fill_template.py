# fill_template.py
from datetime import datetime

from openpyxl import load_workbook
from database import SessionLocal
import models
import os
from collections import defaultdict
from sqlalchemy.orm import joinedload


def fill_template(orders, packaging_type="Упаковка"):
    """Заполняет вертикальный шаблон (товары в строках, точки в колонках)."""
    filepath = os.path.join("exports", "template.xlsx")
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"{filepath} не найден. Сначала нужно сгенерировать шаблон.")

    wb = load_workbook(filepath)
    ws = wb.active

    # --- Карта точек: имя -> колонка ---
    point_col_map = {}
    for col in range(3, ws.max_column + 1):
        point_name = ws.cell(row=1, column=col).value
        if point_name:
            point_col_map[point_name.strip().lower()] = col  # без регистра

    # --- Карта subtypes: subtype -> строка ---
    subtype_row_map = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        unit = ws.cell(row=row, column=2).value
        # subtype — это строка, где колонка B пуста
        if name and unit is None:
            subtype_row_map[name.strip().lower()] = row

    # --- Сбор данных по заказам ---
    data_map = defaultdict(int)
    for order in orders:
        point_name = order.point.name.strip().lower() if order.point else None
        if not point_name:
            continue
        for item in order.items:
            if item.type.lower() != packaging_type.lower():  # учитываем тип из параметра
                continue
            key = (item.subtype.strip().lower() if item.subtype else "",
                   point_name,
                   item.name.strip().lower())
            data_map[key] += item.count

    # --- Запись данных ---
    for (subtype, point_name, item_name), count in data_map.items():
        if subtype not in subtype_row_map and subtype != 'unsub':
            print(f"Subtype '{subtype}' не найден в шаблоне")
            continue
        if point_name not in point_col_map:
            print(f"Точка '{point_name}' не найдена в шаблоне")
            continue

        try:
            subtype_row = subtype_row_map[subtype]
        except:
            subtype_row = 2
        row_to_write = None

        for row in range(subtype_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value and cell.value.strip().lower() == item_name:
                row_to_write = row
                break

        if not row_to_write:
            for row in range(subtype_row + 1, ws.max_row + 1):
                if ws.cell(row=row, column=1).value is None:
                    ws.cell(row=row, column=1, value=item_name)
                    row_to_write = row
                    break

        if not row_to_write:
            print(f"Не удалось найти или создать строку для товара '{item_name}'")
            continue

        col = point_col_map[point_name]
        ws.cell(row=row_to_write, column=col, value=count)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"filled_template_{packaging_type}_{timestamp}.xlsx"
    filled_path = os.path.join("exports", filename)
    wb.save(filled_path)
    print(f"✅ Вертикальный шаблон успешно заполнен: {filled_path}")
    return filled_path


def fill_template_horizontal(orders, packaging_type="Упаковка"):
    """Заполняет горизонтальный шаблон (точки в строках, товары в колонках)."""
    filepath = os.path.join("exports", "template_horizontal.xlsx")
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"{filepath} не найден. Сначала нужно сгенерировать горизонтальный шаблон.")

    wb = load_workbook(filepath)
    ws = wb.active

    # --- Карта товаров: имя -> колонка ---
    product_col_map = {}
    for col in range(2, ws.max_column + 1):
        product_name = ws.cell(row=1, column=col).value
        if product_name:
            product_col_map[product_name.strip().lower()] = col

    # --- Карта точек: имя -> строка ---
    point_row_map = {}
    for row in range(2, ws.max_row + 1):
        point_name = ws.cell(row=row, column=1).value
        if point_name:
            point_row_map[point_name.strip().lower()] = row

    # --- Сбор данных ---
    data_map = defaultdict(int)
    for order in orders:
        point_name = order.point.name.strip().lower() if order.point else None
        if not point_name:
            continue
        for item in order.items:
            if item.type.lower() != packaging_type.lower():
                continue
            item_name = item.name.strip().lower()
            data_map[(point_name, item_name)] += item.count

    # --- Запись данных ---
    for (point_name, item_name), count in data_map.items():
        if point_name not in point_row_map:
            print(f"⚠️ Точка '{point_name}' не найдена в шаблоне")
            continue
        if item_name not in product_col_map:
            print(f"⚠️ Товар '{item_name}' не найден в шаблоне")
            continue

        row = point_row_map[point_name]
        col = product_col_map[item_name]
        ws.cell(row=row, column=col, value=count)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"filled_template_{packaging_type}_{timestamp}.xlsx"
    filled_path = os.path.join("exports", filename)
    wb.save(filled_path)
    print(f"✅ Горизонтальный шаблон успешно заполнен: {filled_path}")
    return filled_path


# ===== Пример использования =====
if __name__ == "__main__":
    db = SessionLocal()
    try:
        orders = db.query(models.Order).options(
            joinedload(models.Order.point),
            joinedload(models.Order.items)
        ).all()
    finally:
        db.close()

    # Пример вызова вертикального шаблона
    fill_template(orders, packaging_type="Часть продуктов")

    # Пример вызова горизонтального шаблона
    # fill_template_horizontal(orders, packaging_type="Упаковка")
