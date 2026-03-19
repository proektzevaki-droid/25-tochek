from fastapi import FastAPI, Depends, HTTPException, Request, Body, Query
from typing import Union, Any, List, Optional
from pydantic import ValidationError
import logging
from fastapi.responses import FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from sqlalchemy.orm import Session
from sqlalchemy.orm import joinedload
from database import SessionLocal, engine, Base
import models, schemas
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from models import moscow_now
import os
from fill_template import fill_template, fill_template_horizontal
from generate_template import generate_template, generate_template_horizontal
from fastapi.staticfiles import StaticFiles
from models import OrderItem, Order  # ваша модель
from typing import List
from dotenv import load_dotenv
import requests
import base64
from io import BytesIO
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
from email.utils import formatdate, make_msgid
from urllib.parse import quote

Base.metadata.create_all(bind=engine)

# Загружаем переменные окружения из .env файла
load_dotenv()

app = FastAPI(title="Dashboard API")
templates = Jinja2Templates(directory="templates")

# Инициализация rate limiter
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.mount("/static", StaticFiles(directory="static"), name="static")

# Получаем настройки из переменных окружения
N8N_WEBHOOK_URL = os.getenv("N8N_WEBHOOK_URL", "")
BASE_URL = os.getenv("BASE_URL", "http://77.73.232.184:12000")
EXPORT_TEMP_DIR = os.getenv("EXPORT_TEMP_DIR", "exports/temp")

# SMTP настройки для отправки email
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM_EMAIL = os.getenv("SMTP_FROM_EMAIL", "")
SMTP_FROM_NAME = os.getenv("SMTP_FROM_NAME", "")

# Создаем директорию для временных файлов, если её нет
os.makedirs(EXPORT_TEMP_DIR, exist_ok=True)

# Настройка логирования
logger = logging.getLogger(__name__)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ---- Dashboard ----
@app.get("/dashboard")
def dashboard(request: Request):
    return templates.TemplateResponse("dashboard.html", {"request": request})

@app.get("/dashboard-v2")
def dashboard_v2(request: Request):
    return templates.TemplateResponse("dashboard-v2.html", {"request": request})

# ---- Points ----
@app.get("/points", response_model=list[schemas.Point])
def read_points(db: Session = Depends(get_db)):
    # Фильтруем удаленные точки - не выводим их на фронт
    return db.query(models.Point).filter(models.Point.is_deleted == False).all()

@app.post("/points", response_model=schemas.Point)
def create_point(point: schemas.PointCreate, db: Session = Depends(get_db)):
    db_point = models.Point(**point.dict())
    db.add(db_point)
    db.commit()
    db.refresh(db_point)
    return db_point

@app.patch("/points/{point_id}", response_model=schemas.Point)
def update_point(point_id: int, point_update: schemas.PointUpdate, db: Session = Depends(get_db)):
    # Обновление точки продаж
    db_point = db.query(models.Point).filter(models.Point.id == point_id).first()
    if not db_point:
        raise HTTPException(status_code=404, detail="Point not found")
    
    # Обновляем только переданные поля
    update_data = point_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_point, key, value)
    
    db.commit()
    db.refresh(db_point)
    return db_point

@app.patch("/points/delete")
def soft_delete_points(ids: list[int] = Body(...), db: Session = Depends(get_db)):
    # Мягкое удаление - устанавливаем is_deleted = True
    db.query(models.Point).filter(models.Point.id.in_(ids)).update(
        {models.Point.is_deleted: True}, synchronize_session=False
    )
    db.commit()
    return {"message": "Points marked as deleted"}


# 1️⃣ — Получить все подтверждённые записи истории
@app.get("/history/confirmed", response_model=list[schemas.History])
def get_confirmed_history(db: Session = Depends(get_db)):
    confirmed_records = db.query(models.History).filter(models.History.status == "принята").all()
    return confirmed_records


# 2️⃣ — Подтвердить запись по ID
@app.post("/history/{history_id}/confirm", response_model=schemas.History)
def confirm_history(history_id: int, db: Session = Depends(get_db)):
    history_record = db.query(models.History).filter(models.History.id == history_id).first()
    if not history_record:
        raise HTTPException(status_code=404, detail="История не найдена")

    history_record.status = "принята"
    db.add(history_record)
    db.commit()
    db.refresh(history_record)
    return history_record


# 3️⃣ — Скачать файл истории по ID
@app.get("/history/{history_id}/download")
def download_history_file(history_id: int, db: Session = Depends(get_db)):
    history_record = db.query(models.History).filter(models.History.id == history_id).first()
    if not history_record:
        raise HTTPException(status_code=404, detail="Файл истории не найден")

    return  FileResponse(
        history_record.file_path,
        filename=history_record.file_path.split("/")[-1],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---- Products ----
@app.get("/products", response_model=list[schemas.Product])
def read_products(db: Session = Depends(get_db)):
    """Получить список всех товаров"""
    return db.query(models.Product).all()

@app.post("/products", response_model=schemas.Product)
def create_product(product: schemas.ProductCreate, db: Session = Depends(get_db)):
    """Создать новый товар"""
    db_product = models.Product(**product.dict())
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    return db_product

@app.patch("/products/{product_id}", response_model=schemas.Product)
def update_product(product_id: int, product_update: schemas.ProductUpdate, db: Session = Depends(get_db)):
    """Обновить товар"""
    db_product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Обновляем только переданные поля
    update_data = product_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_product, key, value)
    
    db.commit()
    db.refresh(db_product)
    return db_product

@app.delete("/products/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    """Удалить товар"""
    db_product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    db.delete(db_product)
    db.commit()
    return {"message": "Product deleted"}

@app.get("/products/types")
def get_product_types(db: Session = Depends(get_db)):
    """Получить список всех уникальных значений product_type из таблицы product"""
    # Получаем уникальные значения product_type, исключая None и пустые строки
    types = db.query(models.Product.product_type).filter(
        models.Product.product_type.isnot(None),
        models.Product.product_type != ""
    ).distinct().all()
    
    # Преобразуем результат в список строк
    product_types = [t[0] for t in types if t[0]]
    
    return {"product_types": sorted(product_types)}

@app.get("/export/by-type")
@limiter.limit("5/second")
def export_by_type(request: Request, product_type: str = Query(..., description="Тип товара для экспорта"), db: Session = Depends(get_db)):
    """
    Универсальный endpoint для экспорта заказов по типу товара.
    Генерирует и заполняет шаблон на основе подтверждённых заказов с указанным типом товара.
    """
    # 1️⃣ — Загружаем заказы
    orders = (
        db.query(models.Order)
        .options(
            joinedload(models.Order.point),
            joinedload(models.Order.items)
        )
        .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
        .all()
    )
    
    # 2️⃣ — Фильтруем только товары указанного типа
    for order in orders:
        order.items = [i for i in order.items if i.type and i.type.lower() == product_type.lower()]
    
    # Удаляем заказы без элементов после фильтрации
    orders = [o for o in orders if o.items]
    
    if not orders:
        return {"message": "Нет данных для заполнения шаблона"}
    
    # 3️⃣ — Создаём шаблон (используем unsub как подтип по умолчанию)
    subtypes = {
        "unsub": {
            "unit": "шт.",
            "items": []
        }
    }
    
    # Собираем уникальные названия товаров из заказов
    product_names = set()
    for order in orders:
        for item in order.items:
            if item.name:
                product_names.add(item.name.lower())
    
    subtypes["unsub"]["items"] = sorted(list(product_names))
    
    generate_template(subtypes)
    
    # 4️⃣ — Заполняем шаблон
    filled_path = fill_template(orders, packaging_type=product_type)
    
    order_item_ids = [order.id for order in orders]
    
    # 5️⃣ — Создаём запись в истории
    history_record = models.History(
        file_path=filled_path,
        created_at=moscow_now()
    )
    db.add(history_record)
    db.commit()
    
    # 6️⃣ — Возвращаем файл
    filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = FileResponse(
        filled_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
    response.headers["X-History-ID"] = str(history_record.id)
    return response

# ---- Suppliers ----
@app.get("/suppliers/simple", response_model=list[schemas.SupplierSimple])
def get_suppliers_simple(db: Session = Depends(get_db)):
    """Получить список поставщиков с минимальной информацией: id, name и доступные способы связи"""
    suppliers = db.query(models.Supplier).all()
    result = []
    for supplier in suppliers:
        available_contacts = []
        if supplier.email:
            available_contacts.append('email')
        if supplier.telegram or supplier.telegram_id:
            available_contacts.append('telegram')
        
        result.append({
            "id": supplier.id,
            "name": supplier.name,
            "available_contacts": available_contacts
        })
    return result

@app.get("/suppliers", response_model=list[schemas.SupplierResponse])
def read_suppliers(db: Session = Depends(get_db)):
    """Получить список всех поставщиков с привязанными товарами и правилами"""
    suppliers = db.query(models.Supplier).options(
        joinedload(models.Supplier.products).joinedload(models.SupplierProduct.product),
        joinedload(models.Supplier.rules)
    ).all()
    result = []
    for supplier in suppliers:
        supplier_dict = {
            "id": supplier.id,
            "name": supplier.name,
            "email": supplier.email,
            "telegram": supplier.telegram,
            "has_telegram_id": supplier.telegram_id is not None,  # Булево значение вместо telegram_id
            "products": [
                {
                    "id": sp.product.id,
                    "name": sp.product.name,
                    "unit": sp.product.unit,
                    "product_type": sp.product.product_type
                }
                for sp in supplier.products
            ],
            "rules": [
                {
                    "id": rule.id,
                    "rule_text": rule.rule_text,
                    "priority": rule.priority,
                    "is_active": rule.is_active,
                    "supplier_id": rule.supplier_id
                }
                for rule in supplier.rules
            ]
        }
        result.append(supplier_dict)
    return result

@app.post("/suppliers", response_model=schemas.Supplier)
def create_supplier(supplier: schemas.SupplierCreate, db: Session = Depends(get_db)):
    """Добавить нового поставщика"""
    db_supplier = models.Supplier(**supplier.dict())
    db.add(db_supplier)
    db.commit()
    db.refresh(db_supplier)
    return db_supplier

@app.patch("/suppliers/{supplier_id}", response_model=schemas.Supplier)
def update_supplier(supplier_id: int, supplier_update: schemas.SupplierUpdate, db: Session = Depends(get_db)):
    """Обновить поставщика"""
    db_supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not db_supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Обновляем только переданные поля
    update_data = supplier_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_supplier, key, value)
    
    db.commit()
    
    # Перезагружаем с связанными данными после обновления
    db_supplier = db.query(models.Supplier).options(
        joinedload(models.Supplier.products).joinedload(models.SupplierProduct.product),
        joinedload(models.Supplier.rules)
    ).filter(models.Supplier.id == supplier_id).first()
    
    # Формируем ответ вручную, как в read_suppliers
    supplier_dict = {
        "id": db_supplier.id,
        "name": db_supplier.name,
        "email": db_supplier.email,
        "telegram": db_supplier.telegram,
        "telegram_id": db_supplier.telegram_id,
        "products": [
            {
                "id": sp.product.id,
                "name": sp.product.name,
                "unit": sp.product.unit,
                "product_type": sp.product.product_type
            }
            for sp in db_supplier.products
        ],
        "rules": [
            {
                "id": r.id,
                "rule_text": r.rule_text,
                "priority": r.priority,
                "is_active": r.is_active,
                "supplier_id": r.supplier_id
            }
            for r in db_supplier.rules
        ]
    }
    return supplier_dict

@app.delete("/suppliers/{supplier_id}")
def delete_supplier(supplier_id: int, db: Session = Depends(get_db)):
    """Удалить поставщика по ID"""
    db_supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not db_supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    db.delete(db_supplier)
    db.commit()
    return {"message": "Supplier deleted"}

@app.post("/suppliers/update-telegram-id")
def update_supplier_telegram_id(
    telegram_id: int = Body(..., description="Telegram ID пользователя"),
    telegram_username: str = Body(..., description="Telegram username (без @)"),
    db: Session = Depends(get_db)
):
    """
    Обновить telegram_id для поставщика по его telegram username.
    Используется при первом взаимодействии с ботом.
    
    Принимает:
    - telegram_id: числовой ID из message.from.id или message.chat.id
    - telegram_username: username из message.from.username (без символа @)
    
    Находит поставщика по полю telegram (может быть в формате @username или username)
    и обновляет telegram_id.
    """
    # Нормализуем username: убираем @ если есть
    normalized_username = telegram_username.lstrip('@').lower()
    
    # Ищем поставщика по telegram полю
    # Может быть в формате @username, username или просто совпадать
    suppliers = db.query(models.Supplier).filter(
        models.Supplier.telegram.isnot(None)
    ).all()
    
    supplier = None
    for s in suppliers:
        if s.telegram:
            # Нормализуем telegram поле поставщика
            supplier_telegram = s.telegram.lstrip('@').lower()
            if supplier_telegram == normalized_username:
                supplier = s
                break
    
    if not supplier:
        raise HTTPException(
            status_code=404, 
            detail=f"Поставщик с telegram username '{telegram_username}' не найден"
        )
    
    # Обновляем telegram_id
    supplier.telegram_id = telegram_id
    db.commit()
    db.refresh(supplier)
    
    return {
        "success": True,
        "message": f"Telegram ID обновлен для поставщика {supplier.name}",
        "supplier_id": supplier.id,
        "supplier_name": supplier.name,
        "telegram_id": supplier.telegram_id
    }

# ---- Supplier Products ----
@app.get("/suppliers/{supplier_id}/available-products", response_model=list[schemas.Product])
def get_available_products(supplier_id: int, db: Session = Depends(get_db)):
    """Получить список товаров, которые еще не привязаны к поставщику"""
    # Получаем ID всех товаров, уже привязанных к этому поставщику
    linked_product_ids = [
        row[0] for row in db.query(models.SupplierProduct.product_id).filter(
            models.SupplierProduct.supplier_id == supplier_id
        ).all()
    ]
    
    # Получаем все товары, которые не привязаны к поставщику
    if linked_product_ids:
        available_products = db.query(models.Product).filter(
            ~models.Product.id.in_(linked_product_ids)
        ).all()
    else:
        # Если нет привязанных товаров, возвращаем все товары
        available_products = db.query(models.Product).all()
    
    return available_products

@app.post("/suppliers/{supplier_id}/products/{product_id}")
def add_product_to_supplier(supplier_id: int, product_id: int, db: Session = Depends(get_db)):
    """Привязать товар к поставщику"""
    # Проверяем существование поставщика
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Проверяем существование товара
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Проверяем, не привязан ли уже товар к поставщику
    existing = db.query(models.SupplierProduct).filter(
        models.SupplierProduct.supplier_id == supplier_id,
        models.SupplierProduct.product_id == product_id
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="Product already linked to this supplier")
    
    # Создаем связь
    supplier_product = models.SupplierProduct(
        supplier_id=supplier_id,
        product_id=product_id
    )
    db.add(supplier_product)
    db.commit()
    db.refresh(supplier_product)
    
    return {"message": "Product linked to supplier", "supplier_product_id": supplier_product.id}

@app.delete("/suppliers/{supplier_id}/products/{product_id}")
def remove_product_from_supplier(supplier_id: int, product_id: int, db: Session = Depends(get_db)):
    """Отвязать товар от поставщика"""
    supplier_product = db.query(models.SupplierProduct).filter(
        models.SupplierProduct.supplier_id == supplier_id,
        models.SupplierProduct.product_id == product_id
    ).first()
    
    if not supplier_product:
        raise HTTPException(status_code=404, detail="Product not linked to this supplier")
    
    db.delete(supplier_product)
    db.commit()
    
    return {"message": "Product unlinked from supplier"}

@app.get("/suppliers/{supplier_id}/rules", response_model=list[schemas.SupplierRule])
def get_supplier_rules(supplier_id: int, db: Session = Depends(get_db)):
    """Получить все правила поставщика"""
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    rules = db.query(models.SupplierRule).filter(
        models.SupplierRule.supplier_id == supplier_id
    ).order_by(models.SupplierRule.priority.asc()).all()
    return rules

@app.post("/suppliers/{supplier_id}/rules", response_model=schemas.SupplierRule)
def create_supplier_rule(supplier_id: int, rule: schemas.SupplierRuleCreate, db: Session = Depends(get_db)):
    """Добавить новое правило для поставщика"""
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    db_rule = models.SupplierRule(**rule.dict(), supplier_id=supplier_id)
    db.add(db_rule)
    db.commit()
    db.refresh(db_rule)
    return db_rule

@app.patch("/suppliers/{supplier_id}/rules/{rule_id}", response_model=schemas.SupplierRule)
def update_supplier_rule(supplier_id: int, rule_id: int, rule_update: schemas.SupplierRuleUpdate, db: Session = Depends(get_db)):
    """Обновить правило поставщика"""
    db_rule = db.query(models.SupplierRule).filter(
        models.SupplierRule.id == rule_id,
        models.SupplierRule.supplier_id == supplier_id
    ).first()
    
    if not db_rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    
    update_data = rule_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_rule, field, value)
    
    db.commit()
    db.refresh(db_rule)
    return db_rule

@app.delete("/suppliers/{supplier_id}/rules/{rule_id}")
def delete_supplier_rule(supplier_id: int, rule_id: int, db: Session = Depends(get_db)):
    """Удалить правило поставщика"""
    db_rule = db.query(models.SupplierRule).filter(
        models.SupplierRule.id == rule_id,
        models.SupplierRule.supplier_id == supplier_id
    ).first()
    
    if not db_rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    
    db.delete(db_rule)
    db.commit()
    return {"message": "Rule deleted"}

@app.get("/suppliers/export")
def export_suppliers_data(db: Session = Depends(get_db)):
    """Получить данные о поставщиках, их товарах и исключениях для нейросети
    
    Возвращает JSON в формате:
    {
        "Поставщики": {
            "Поставщик 1": {
                "id": 1,
                "products": [
                    {"id": 1, "name": "Товар 1", "product_type": "Тип товара", "unit": "шт."},
                    {"id": 2, "name": "Товар 2", "product_type": "Тип товара", "unit": "кг"}
                ]
            },
            "Поставщик 2": {
                "id": 2,
                "products": [
                    {"id": 3, "name": "Товар 3", "product_type": "Тип товара", "unit": "л"}
                ]
            }
        },
        "Исключения": [
            {
                "rule_text": "Текст правила",
                "priority": 100,
                "supplier_id": 3,
                "supplier_name": "ПРОТЕК"
            }
        ]
    }
    """
    # Получаем всех поставщиков с товарами
    suppliers = db.query(models.Supplier).options(
        joinedload(models.Supplier.products).joinedload(models.SupplierProduct.product)
    ).all()
    
    # Создаем словарь для быстрого доступа к поставщикам по ID
    suppliers_by_id = {s.id: s for s in suppliers}
    
    # Формируем объект поставщиков: ключ - название, значение - объект с id и товарами
    suppliers_dict = {}
    for supplier in suppliers:
        # Получаем товары поставщика с ID, названием, типом и единицей измерения
        products = [
            {
                "id": sp.product.id,
                "name": sp.product.name,
                "product_type": sp.product.product_type,
                "unit": sp.product.unit
            }
            for sp in supplier.products
        ]
        suppliers_dict[supplier.name] = {
            "id": supplier.id,
            "products": products
        }
    
    # Получаем все активные правила (исключения) из таблицы supplier_rules
    # Запрос напрямую к таблице supplier_rules с фильтром по is_active
    all_rules = db.query(models.SupplierRule).filter(
        models.SupplierRule.is_active == True
    ).order_by(models.SupplierRule.priority.asc()).all()
    
    # Формируем список исключений с полной информацией
    exceptions = []
    for rule in all_rules:
        supplier = suppliers_by_id.get(rule.supplier_id)
        exceptions.append({
            "rule_text": rule.rule_text,
            "priority": rule.priority or 100,
            "supplier_id": rule.supplier_id,
            "supplier_name": supplier.name if supplier else None
        })
    
    return {
        "Поставщики": suppliers_dict,
        "Исключения": exceptions
    }

# ---- OrderItems ----
@app.put("/order-items/update-status", response_model=dict)
def update_order_items_status(item_ids: List[int], db: Session = Depends(get_db)):
    if not item_ids:
        raise HTTPException(status_code=400, detail="List of IDs cannot be empty")

    query = db.query(Order).filter(Order.id.in_(item_ids))
    items = query.all()
    if not items:
        raise HTTPException(status_code=404, detail="No items found for provided IDs")

    # Обновляем статус всех найденных объектов
    for item in items:
        item.sent_to_supplier = True

    db.commit()

    return {"updated_ids": [item.id for item in items], "status": True}

@app.patch("/order-items/update-supplier", response_model=dict)
def update_order_items_supplier(
    product_id: int = Query(..., description="ID товара"),
    old_supplier_id: int = Query(..., description="Старый ID поставщика"),
    new_supplier_id: int = Query(..., description="Новый ID поставщика"),
    db: Session = Depends(get_db)
):
    """Обновить поставщика для всех order_items с указанным product_id и old_supplier_id"""
    # Находим все order_items с указанным product_id и old_supplier_id
    items = db.query(models.OrderItem).filter(
        models.OrderItem.product_id == product_id,
        models.OrderItem.suppliers_id == old_supplier_id
    ).all()
    
    if not items:
        raise HTTPException(status_code=404, detail="No order items found with specified product_id and supplier_id")
    
    # Обновляем suppliers_id для всех найденных элементов
    updated_count = 0
    for item in items:
        item.suppliers_id = new_supplier_id
        updated_count += 1
    
    db.commit()
    
    return {
        "updated_count": updated_count,
        "product_id": product_id,
        "old_supplier_id": old_supplier_id,
        "new_supplier_id": new_supplier_id
    }

@app.patch("/order-items/update-supplier-point", response_model=dict)
def update_order_items_supplier_point(
    product_id: int = Query(..., description="ID товара"),
    point_id: int = Query(..., description="ID точки"),
    old_supplier_id: int = Query(..., description="Старый ID поставщика"),
    new_supplier_id: int = Query(..., description="Новый ID поставщика"),
    db: Session = Depends(get_db)
):
    """Обновить поставщика для order_items с указанным product_id, point_id и old_supplier_id"""
    # Находим заказы для указанной точки
    orders = db.query(models.Order).filter(
        models.Order.point_id == point_id,
        models.Order.status == "confirmed",
        models.Order.sent_to_supplier == False
    ).all()
    
    if not orders:
        raise HTTPException(status_code=404, detail="No orders found for specified point_id")
    
    # Находим все order_items с указанным product_id, old_supplier_id в этих заказах
    order_ids = [order.id for order in orders]
    items = db.query(models.OrderItem).filter(
        models.OrderItem.order_id.in_(order_ids),
        models.OrderItem.product_id == product_id,
        models.OrderItem.suppliers_id == old_supplier_id,
        models.OrderItem.is_sent_to_supplier == False
    ).all()
    
    if not items:
        raise HTTPException(status_code=404, detail="No order items found with specified product_id, point_id and supplier_id")
    
    # Обновляем suppliers_id для всех найденных элементов
    updated_count = 0
    for item in items:
        item.suppliers_id = new_supplier_id
        updated_count += 1
    
    db.commit()
    
    return {
        "updated_count": updated_count,
        "product_id": product_id,
        "point_id": point_id,
        "old_supplier_id": old_supplier_id,
        "new_supplier_id": new_supplier_id
    }

@app.patch("/order-items/update-count", response_model=dict)
def update_order_items_count(
    product_id: int = Body(..., description="ID товара"),
    point_id: int = Body(..., description="ID точки"),
    product_type: str = Body(..., description="Тип товара"),
    new_count: float = Body(..., description="Новое общее количество"),
    db: Session = Depends(get_db)
):
    """Обновить количество для order_items с указанным product_id, point_id и типом товара.
    
    Обновляет все order_items для данной точки и товара так, чтобы их сумма была равна new_count.
    Если есть несколько заказов, количество распределяется равномерно между ними.
    """
    # Находим заказы для указанной точки
    orders = db.query(models.Order).filter(
        models.Order.point_id == point_id,
        models.Order.status == "confirmed",
        models.Order.sent_to_supplier == False
    ).all()
    
    if not orders:
        raise HTTPException(status_code=404, detail="No orders found for specified point_id")
    
    # Находим все order_items с указанным product_id и типом в этих заказах
    order_ids = [order.id for order in orders]
    items = db.query(models.OrderItem).filter(
        models.OrderItem.order_id.in_(order_ids),
        models.OrderItem.product_id == product_id,
        models.OrderItem.type.isnot(None),
        models.OrderItem.type.ilike(f"%{product_type}%"),
        models.OrderItem.is_sent_to_supplier == False
    ).all()
    
    if not items:
        raise HTTPException(status_code=404, detail="No order items found with specified product_id, point_id and product_type")
    
    # Если количество = 0 или меньше, устанавливаем None для всех элементов
    if new_count <= 0:
        for item in items:
            item.count = None
    else:
        # Распределяем количество равномерно между всеми заказами
        # Если заказов несколько, каждый получает равную долю
        count_per_item = new_count / len(items) if items else 0
        
        for item in items:
            item.count = count_per_item
    
    db.commit()
    
    return {
        "success": True,
        "updated_count": len(items),
        "product_id": product_id,
        "point_id": point_id,
        "product_type": product_type,
        "new_count": new_count
    }

def get_owner_fill_color(owner: Optional[str]) -> Optional[PatternFill]:
    """
    Возвращает цвет заливки для ячейки заголовка в зависимости от owner.
    ИП Афонин: #72ac50
    ИП Солдатов: без цвета (None)
    """
    if not owner:
        return None
    
    owner_lower = owner.lower().strip()
    if "афонин" in owner_lower:
        return PatternFill(start_color="72ac50", end_color="72ac50", fill_type="solid")
    
    return None

def apply_excel_styles(ws, points_list, products_data, supplier_items=None):
    """
    Применяет стили к Excel таблице: границы, заголовки, четные строки, автоширина первого столбца.
    
    Args:
        ws: Worksheet объект openpyxl
        points_list: Список точек [(point_id, point_name, point_owner), ...]
        products_data: Словарь с данными о товарах
        supplier_items: Опциональный список элементов для расчета максимальной длины названия
    """
    # Определяем стили
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    header_fill_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    even_row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Устанавливаем высоту строки заголовка
    ws.row_dimensions[1].height = 90
    
    # Рассчитываем максимальную длину названия товара для автоширины первого столбца
    max_product_name_length = 0
    if supplier_items:
        for item in supplier_items:
            product_name = f"{item.get('product_name', '')}"
            if item.get('product_unit'):
                product_name += f", {item.get('product_unit', '')}"
            max_product_name_length = max(max_product_name_length, len(product_name))
    else:
        for product_id, data in products_data.items():
            product_name = f"{data.get('name', '')}"
            if data.get('unit'):
                product_name += f", {data.get('unit', '')}"
            max_product_name_length = max(max_product_name_length, len(product_name))
    
    # Устанавливаем ширину первого столбца
    ws.column_dimensions['A'].width = max(max_product_name_length * 1.2 + 3, 15)
    
    # Применяем стили к заголовкам
    # Первая ячейка заголовка "Товар"
    header_cell_1 = ws.cell(row=1, column=1)
    header_cell_1.fill = header_fill_gray
    header_cell_1.font = header_font
    header_cell_1.alignment = Alignment(vertical='center', horizontal='center')
    header_cell_1.border = thin_border
    
    # Остальные ячейки заголовка (точки)
    for col_idx, (point_id, point_name, point_owner) in enumerate(points_list, start=2):
        cell = ws.cell(row=1, column=col_idx)
        fill_color = get_owner_fill_color(point_owner)
        if fill_color:
            cell.fill = fill_color
        else:
            cell.fill = header_fill_gray
        cell.font = header_font
        cell.alignment = Alignment(text_rotation=90, vertical='center', horizontal='center', wrap_text=True)
        cell.border = thin_border
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15
    
    # Применяем стили к строкам данных
    row_num = 2
    for product_id, data in sorted(products_data.items(), key=lambda x: x[1].get("name", "")):
        is_even_row = (row_num % 2 == 0)
        num_cols = len(points_list) + 1  # +1 для колонки "Товар"
        
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if is_even_row:
                cell.fill = even_row_fill
            if col_idx == 1:
                cell.alignment = Alignment(vertical='center', horizontal='left')
            else:
                cell.alignment = Alignment(vertical='center', horizontal='center')
        
        row_num += 1

@app.get("/export/supplier-excel")
@limiter.limit("5/second")
def generate_supplier_excel(
    request: Request,
    supplier_id: int = Query(..., description="ID поставщика"),
    product_type: str = Query(..., description="Тип продукта"),
    db: Session = Depends(get_db)
):
    """
    Генерирует Excel файл для конкретного поставщика с товарами указанного типа.
    Возвращает файл для скачивания.
    """
    # Получаем заказы со статусом confirmed и sent_to_supplier = False
    orders = (
        db.query(models.Order)
        .options(
            joinedload(models.Order.items).joinedload(models.OrderItem.product),
            joinedload(models.Order.point)
        )
        .filter(
            models.Order.status == "confirmed",
            models.Order.sent_to_supplier == False
        )
        .all()
    )
    
    # Фильтруем товары: только для указанного поставщика и типа продукта, которые еще не отправлены
    supplier_items = []
    points_set = set()
    products_map = {}
    
    for order in orders:
        if not order.point:
            continue
            
        for item in order.items:
            if (item.suppliers_id == supplier_id and 
                item.type and 
                item.type.lower() == product_type.lower() and
                not item.is_sent_to_supplier):
                point_owner = order.point.owner if order.point.owner else ""
                supplier_items.append({
                    "order_id": order.id,
                    "point_id": order.point.id,
                    "point_name": order.point.name,
                    "point_owner": point_owner,
                    "product_id": item.product_id,
                    "product_name": item.name or (item.product.name if item.product else ""),
                    "product_unit": item.product.unit if item.product else "",
                    "count": item.count or 0
                })
                points_set.add((order.point.id, order.point.name, point_owner))
                if item.product_id and item.product_id not in products_map:
                    products_map[item.product_id] = {
                        "name": item.name or (item.product.name if item.product else ""),
                        "unit": item.product.unit if item.product else ""
                    }
    
    if not supplier_items:
        raise HTTPException(status_code=404, detail="Нет товаров для данного поставщика и типа продукта")
    
    # Получаем информацию о поставщике
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Поставщик не найден")
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Выгрузка"
    
    # Заголовки: Товар, точки (с owner)
    header_row = ["Товар"]
    # Сортируем по owner, затем по названию
    points_list = sorted(list(points_set), key=lambda x: (x[2] or "", x[1]))
    logger = logging.getLogger(__name__)
    logger.info(f"[SUPPLIER EXCEL] Points list: {points_list}")
    for point_id, point_name, point_owner in points_list:
        if point_owner:
            header_text = f"{point_name} ({point_owner})"
        else:
            header_text = point_name
        header_row.append(header_text)
    ws.append(header_row)
    
    # Группируем данные по товарам
    products_data = {}
    for item in supplier_items:
        product_key = item["product_id"]
        if product_key not in products_data:
            products_data[product_key] = {
                "name": item["product_name"],
                "unit": item["product_unit"],
                "points": {}
            }
        
        point_id = item["point_id"]
        if point_id not in products_data[product_key]["points"]:
            products_data[product_key]["points"][point_id] = 0
        products_data[product_key]["points"][point_id] += item["count"]
    
    # Заполняем строки данных
    for product_id, data in sorted(products_data.items(), key=lambda x: x[1]["name"]):
        row = []
        product_name = f"{data['name']}"
        if data['unit']:
            product_name += f", {data['unit']}"
        row.append(product_name)
        
        for point_id, point_name, point_owner in points_list:
            count = data["points"].get(point_id, 0)
            row.append(count if count > 0 else "")
        
        ws.append(row)
    
    # Применяем стили к таблице
    apply_excel_styles(ws, points_list, products_data, supplier_items)
    
    # Сохраняем файл во временную директорию
    filename = f"supplier_{supplier_id}_{product_type}_{moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_TEMP_DIR, filename)
    wb.save(filepath)
    
    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/test/supplier-excel")
@limiter.limit("10/second")
def test_generate_supplier_excel(
    request: Request,
    order_id: Optional[int] = Body(None, description="ID заказа для генерации таблицы"),
    test_data: Optional[dict] = Body(None, description="Тестовые данные для генерации таблицы"),
    db: Session = Depends(get_db)
):
    """
    Тестовая ручка для генерации таблицы поставщика.
    Принимает либо order_id, либо test_data с тестовыми данными.
    Возвращает ссылку на сгенерированный файл.
    """
    # Подготовка данных
    supplier_items = []
    points_set = set()
    products_data = {}
    max_product_name_length = 0
    
    if order_id:
        # Генерируем таблицу на основе реального заказа
        order = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.items).joinedload(models.OrderItem.product),
                joinedload(models.Order.point)
            )
            .filter(models.Order.id == order_id)
            .first()
        )
        
        if not order:
            raise HTTPException(status_code=404, detail="Заказ не найден")
        
        if not order.point:
            raise HTTPException(status_code=400, detail="У заказа не указана точка")
        
        for item in order.items:
            if item.suppliers_id and item.type:
                point_owner = order.point.owner if order.point.owner else ""
                supplier_items.append({
                    "point_id": order.point.id,
                    "point_name": order.point.name,
                    "point_owner": point_owner,
                    "product_id": item.product_id,
                    "product_name": item.name or (item.product.name if item.product else ""),
                    "product_unit": item.product.unit if item.product else "",
                    "count": item.count or 0
                })
                points_set.add((order.point.id, order.point.name, point_owner))
        
        # Добавляем еще несколько тестовых точек для демонстрации
        points_set.add((999, "Тестовая точка 1", "ИП Афонин"))
        points_set.add((998, "Тестовая точка 2", "ИП Солдатов"))
        points_set.add((997, "Тестовая точка 3", None))
        
        # Добавляем тестовые товары
        test_products = [
            {"name": "Товар 1", "unit": "шт.", "count": 10},
            {"name": "Товар 2", "unit": "кг", "count": 5},
            {"name": "Товар 3", "unit": "л", "count": 3},
        ]
        
        for product in test_products:
            products_data[f"test_{product['name']}"] = {
                "name": product['name'],
                "unit": product['unit'],
                "points": {}
            }
            # Распределяем товары по точкам
            for point_id, point_name, point_owner in points_set:
                products_data[f"test_{product['name']}"]["points"][point_id] = product['count']
    
    elif test_data:
        # Генерируем таблицу на основе переданных тестовых данных
        points_list = test_data.get("points", [
            {"id": 1, "name": "Точка 1", "owner": "ИП Афонин"},
            {"id": 2, "name": "Точка 2", "owner": "ИП Солдатов"},
            {"id": 3, "name": "Точка 3", "owner": None}
        ])
        
        products_list = test_data.get("products", [
            {"name": "Товар 1", "unit": "шт."},
            {"name": "Товар 2", "unit": "кг"},
            {"name": "Товар 3", "unit": "л"},
        ])
        
        for point in points_list:
            points_set.add((point.get("id", 0), point.get("name", ""), point.get("owner", "")))
        
        for product in products_list:
            products_data[product.get('name', '')] = {
                "name": product.get('name', ''),
                "unit": product.get('unit', ''),
                "points": {}
            }
            # Заполняем случайными данными
            for point in points_list:
                point_id = point.get("id", 0)
                products_data[product.get('name', '')]["points"][point_id] = test_data.get("counts", {}).get(
                    f"{product.get('name', '')}_{point_id}", 0
                )
    else:
        # Генерируем таблицу с дефолтными тестовыми данными
        points_set.add((1, "Точка 1", "ИП Афонин"))
        points_set.add((2, "Точка 2", "ИП Солдатов"))
        points_set.add((3, "Точка 3", None))
        points_set.add((4, "Точка 4", "ИП Афонин"))
        
        # Выбираем 10 товаров разной длины из реального списка
        test_products = [
            {"name": "Вилка", "unit": "шт."},
            {"name": "Ложка", "unit": "шт."},
            {"name": "Ложка чайная", "unit": "шт."},
            {"name": "Палочка для размешивания", "unit": "шт."},
            {"name": "Стакан 200 мл. прозрачный", "unit": "шт."},
            {"name": "Контейнер (179) прямоугольный 750 мл.", "unit": "шт."},
            {"name": "Крышка для контейнера (179) прямоугольного", "unit": "шт."},
            {"name": "Напиток Фрустайл лимон-лайм ж/б", "unit": "шт."},
            {"name": "Напиток Эвервес без сахара 0,5", "unit": "шт."},
            {"name": "Перчатки виниловые (100 шт/уп)", "unit": "уп."},
        ]
        
        # Распределяем товары по точкам с разными количествами
        # Каждый товар должен быть хотя бы у одной точки
        point_list = sorted(list(points_set), key=lambda x: (x[2] or "", x[1]))
        for idx, product in enumerate(test_products):
            products_data[product['name']] = {
                "name": product['name'],
                "unit": product['unit'],
                "points": {}
            }
            # Заполняем количествами: каждый товар есть минимум у одной точки
            # Используем индекс для разнообразия распределения
            for point_idx, (point_id, point_name, point_owner) in enumerate(point_list):
                # Создаем паттерн распределения: не все товары у всех точек
                # Но каждый товар есть хотя бы у одной точки (гарантируем первую точку)
                if point_idx == 0 or (idx + point_idx) % 3 != 0:
                    # Количество зависит от индекса товара и точки для разнообразия
                    # Используем формулу для получения разных значений от 1 до 20
                    count = ((idx * 7 + point_idx * 5 + 1) % 20) + 1
                    products_data[product['name']]["points"][point_id] = count
    
    if not products_data:
        raise HTTPException(status_code=400, detail="Нет данных для генерации таблицы")
    
    # Логируем для отладки
    logger = logging.getLogger(__name__)
    logger.info(f"[TEST EXCEL] Products count: {len(products_data)}")
    logger.info(f"[TEST EXCEL] Products: {list(products_data.keys())}")
    logger.info(f"[TEST EXCEL] Points count: {len(points_set)}")
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Выгрузка"
    
    # Заголовки
    header_row = ["Товар"]
    points_list = sorted(list(points_set), key=lambda x: (x[2] or "", x[1]))
    for point_id, point_name, point_owner in points_list:
        if point_owner:
            header_text = f"{point_name} ({point_owner})"
        else:
            header_text = point_name
        header_row.append(header_text)
    ws.append(header_row)
    
    # Заполняем строки данных
    for product_id, data in sorted(products_data.items(), key=lambda x: x[1].get("name", "")):
        row = []
        product_name = f"{data.get('name', '')}"
        if data.get('unit'):
            product_name += f", {data.get('unit', '')}"
        row.append(product_name)
        
        for point_id, point_name, point_owner in points_list:
            count = data.get("points", {}).get(point_id, 0)
            row.append(count if count > 0 else "")
        
        ws.append(row)
    
    # Применяем стили к таблице
    apply_excel_styles(ws, points_list, products_data)
    
    # Сохраняем файл
    filename = f"test_supplier_excel_{moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_TEMP_DIR, filename)
    wb.save(filepath)
    
    # Формируем URL для скачивания
    file_url = f"{BASE_URL}/test/supplier-excel/download/{filename}"
    
    return {
        "success": True,
        "message": "Таблица успешно сгенерирована",
        "filename": filename,
        "download_url": file_url,
        "file_path": filepath
    }

@app.get("/test/supplier-excel/download/{filename}")
def download_test_excel(filename: str):
    """
    Скачивание тестового Excel файла по имени файла.
    """
    safe_filename = os.path.basename(filename)
    filepath = os.path.join(EXPORT_TEMP_DIR, safe_filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Файл не найден")
    
    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/send/supplier-telegram")
@limiter.limit("5/second")
def send_supplier_telegram(
    request: Request,
    supplier_id: int = Body(..., description="ID поставщика"),
    product_type: str = Body(..., description="Тип продукта"),
    db: Session = Depends(get_db)
):
    """
    Отправляет заявку поставщику через Telegram через n8n webhook.
    Сначала генерирует Excel файл, затем отправляет данные в n8n.
    """
    if not N8N_WEBHOOK_URL:
        raise HTTPException(status_code=500, detail="N8N_WEBHOOK_URL не настроен в .env файле")
    
    # Получаем информацию о поставщике
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Поставщик не найден")
    
    # Проверяем наличие telegram_id или telegram
    if not supplier.telegram_id and not supplier.telegram:
        raise HTTPException(status_code=400, detail="У поставщика не указан Telegram")
    
    # Генерируем Excel файл
    # Сначала получаем данные для файла
    orders = (
        db.query(models.Order)
        .options(
            joinedload(models.Order.items).joinedload(models.OrderItem.product),
            joinedload(models.Order.point)
        )
        .filter(
            models.Order.status == "confirmed",
            models.Order.sent_to_supplier == False
        )
        .all()
    )
    
    supplier_items = []
    points_set = set()
    
    for order in orders:
        if not order.point:
            continue
            
        for item in order.items:
            if (item.suppliers_id == supplier_id and 
                item.type and 
                item.type.lower() == product_type.lower() and
                not item.is_sent_to_supplier):
                supplier_items.append({
                    "order_id": order.id,
                    "point_id": order.point.id,
                    "point_name": order.point.name,
                    "point_owner": order.point.owner,
                    "product_id": item.product_id,
                    "product_name": item.name or (item.product.name if item.product else ""),
                    "product_unit": item.product.unit if item.product else "",
                    "count": item.count or 0,
                    "item_id": item.id
                })
                points_set.add((order.point.id, order.point.name, order.point.owner or ""))
    
    if not supplier_items:
        raise HTTPException(status_code=404, detail="Нет товаров для данного поставщика и типа продукта")
    
    # Создаем Excel файл в памяти
    wb = Workbook()
    ws = wb.active
    ws.title = "Выгрузка"
    
    # Заголовки
    header_row = ["Товар"]
    points_list = sorted(list(points_set), key=lambda x: (x[2] or "", x[1]))
    for point_id, point_name, point_owner in points_list:
        if point_owner:
            header_text = f"{point_name} ({point_owner})"
        else:
            header_text = point_name
        header_row.append(header_text)
    ws.append(header_row)
    
    # Группируем данные по товарам
    products_data = {}
    for item in supplier_items:
        product_key = item["product_id"]
        if product_key not in products_data:
            products_data[product_key] = {
                "name": item["product_name"],
                "unit": item["product_unit"],
                "points": {}
            }
        
        point_id = item["point_id"]
        if point_id not in products_data[product_key]["points"]:
            products_data[product_key]["points"][point_id] = 0
        products_data[product_key]["points"][point_id] += item["count"]
    
    # Заполняем строки данных
    for product_id, data in sorted(products_data.items(), key=lambda x: x[1]["name"]):
        row = []
        product_name = f"{data['name']}"
        if data['unit']:
            product_name += f", {data['unit']}"
        row.append(product_name)
        
        for point_id, point_name, point_owner in points_list:
            count = data["points"].get(point_id, 0)
            row.append(count if count > 0 else "")
        
        ws.append(row)
    
    # Применяем стили к таблице
    apply_excel_styles(ws, points_list, products_data, supplier_items)
    
    # Сохраняем в BytesIO для отправки
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    excel_base64 = base64.b64encode(excel_buffer.read()).decode('utf-8')
    
    # Формируем данные для отправки в n8n
    # Собираем список ID элементов заказа для обновления статуса
    item_ids = [item["item_id"] for item in supplier_items]
    
    # Формируем callback URL для n8n
    callback_url = f"{BASE_URL}/send/supplier-telegram/callback"
    
    payload = {
        "supplier_id": supplier_id,
        "supplier_name": supplier.name,
        "telegram_id": supplier.telegram_id,  # Приоритет telegram_id
        "telegram": supplier.telegram,  # Fallback если telegram_id нет
        "product_type": product_type,
        "excel_file_base64": excel_base64,
        "excel_filename": f"Заявка_{supplier.name}_{product_type}_{moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "item_ids": item_ids,
        "items_count": len(supplier_items),
        "callback_url": callback_url  # URL для отправки результата
    }
    
    # Отправляем запрос в n8n
    try:
        logger = logging.getLogger(__name__)
        
        logger.info(f"Отправка в n8n webhook: {N8N_WEBHOOK_URL}")
        logger.info(f"Payload keys: {list(payload.keys())}")
        
        response = requests.post(N8N_WEBHOOK_URL, json=payload, timeout=30)
        
        logger.info(f"Response status: {response.status_code}")
        logger.info(f"Response text: {response.text[:200]}")
        
        response.raise_for_status()
        
        # Устанавливаем флаг "отправляется" для элементов заказа
        # Фактический статус is_sent_to_supplier будет обновлен через callback от n8n
        for item_id in item_ids:
            item = db.query(models.OrderItem).filter(models.OrderItem.id == item_id).first()
            if item:
                item.is_sending_to_supplier = True
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Заявка отправляется поставщику {supplier.name}. Ожидание подтверждения...",
            "items_count": len(item_ids),
            "item_ids": item_ids,
            "status": "sending"
        }
    except requests.exceptions.HTTPError as e:
        error_detail = f"HTTP {e.response.status_code}: {e.response.text[:200] if e.response else str(e)}"
        logger.error(f"HTTP ошибка при отправке в n8n: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Ошибка при отправке в n8n: {error_detail}")
    except requests.exceptions.RequestException as e:
        logger.error(f"Ошибка при отправке в n8n: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка при отправке в n8n: {str(e)}")

@app.post("/send/supplier-telegram/callback")
def supplier_telegram_callback(
    status: str = Body(..., description="Статус отправки: 'success' или 'error'"),
    item_ids: List[int] = Body(..., description="Список ID элементов заказа для обновления"),
    error_message: Optional[str] = Body(None, description="Сообщение об ошибке (если status='error')"),
    db: Session = Depends(get_db)
):
    """
    Callback endpoint от n8n для подтверждения отправки заявки в Telegram.
    
    Вызывается n8n после успешной или неуспешной отправки сообщения в Telegram.
    """
    if status not in ["success", "error"]:
        raise HTTPException(status_code=400, detail="status должен быть 'success' или 'error'")
    
    if not item_ids:
        raise HTTPException(status_code=400, detail="item_ids не может быть пустым")
    
    # Проверяем, что все item_ids существуют и находятся в статусе "отправляется"
    items = db.query(models.OrderItem).filter(
        models.OrderItem.id.in_(item_ids),
        models.OrderItem.is_sending_to_supplier == True
    ).all()
    
    if len(items) != len(item_ids):
        raise HTTPException(
            status_code=400, 
            detail=f"Не все элементы найдены или не находятся в статусе отправки. Найдено: {len(items)}, ожидалось: {len(item_ids)}"
        )
    
    if status == "success":
        # Успешная отправка - обновляем статус
        for item in items:
            item.is_sent_to_supplier = True
            item.is_sending_to_supplier = False
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Статус обновлен для {len(items)} элементов заказа",
            "items_updated": len(items)
        }
    else:
        # Ошибка отправки - сбрасываем флаг отправки
        for item in items:
            item.is_sending_to_supplier = False
        
        db.commit()
        
        error_msg = error_message or "Ошибка при отправке в Telegram"
        
        return {
            "success": False,
            "message": f"Ошибка отправки: {error_msg}",
            "items_count": len(items),
            "error": error_msg
        }


@app.post("/send/supplier-email")
@limiter.limit("5/second")
def send_supplier_email(
    request: Request,
    supplier_id: int = Body(..., description="ID поставщика"),
    product_type: str = Body(..., description="Тип продукта"),
    db: Session = Depends(get_db)
):
    """
    Отправляет заявку поставщику через Email.
    Сначала генерирует Excel файл, затем отправляет письмо с вложением через SMTP.
    """
    # Проверяем SMTP настройки
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASSWORD or not SMTP_FROM_EMAIL:
        raise HTTPException(status_code=500, detail="SMTP настройки не настроены в .env файле")
    
    # Получаем информацию о поставщике
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Поставщик не найден")
    
    # Проверяем наличие email
    if not supplier.email:
        raise HTTPException(status_code=400, detail="У поставщика не указан email")
    
    # Генерируем Excel файл (та же логика, что и для Telegram)
    orders = (
        db.query(models.Order)
        .options(
            joinedload(models.Order.items).joinedload(models.OrderItem.product),
            joinedload(models.Order.point)
        )
        .filter(
            models.Order.status == "confirmed",
            models.Order.sent_to_supplier == False
        )
        .all()
    )
    
    supplier_items = []
    points_set = set()
    
    for order in orders:
        if not order.point:
            continue
            
        for item in order.items:
            if (item.suppliers_id == supplier_id and 
                item.type and 
                item.type.lower() == product_type.lower() and
                not item.is_sent_to_supplier):
                supplier_items.append({
                    "order_id": order.id,
                    "point_id": order.point.id,
                    "point_name": order.point.name,
                    "point_owner": order.point.owner,
                    "product_id": item.product_id,
                    "product_name": item.name or (item.product.name if item.product else ""),
                    "product_unit": item.product.unit if item.product else "",
                    "count": item.count or 0,
                    "item_id": item.id
                })
                points_set.add((order.point.id, order.point.name, order.point.owner or ""))
    
    if not supplier_items:
        raise HTTPException(status_code=404, detail="Нет товаров для данного поставщика и типа продукта")
    
    # Создаем Excel файл в памяти
    wb = Workbook()
    ws = wb.active
    ws.title = "Выгрузка"
    
    # Заголовки
    header_row = ["Товар"]
    points_list = sorted(list(points_set), key=lambda x: (x[2] or "", x[1]))
    for point_id, point_name, point_owner in points_list:
        if point_owner:
            header_text = f"{point_name} ({point_owner})"
        else:
            header_text = point_name
        header_row.append(header_text)
    ws.append(header_row)
    
    # Группируем данные по товарам
    products_data = {}
    for item in supplier_items:
        product_key = item["product_id"]
        if product_key not in products_data:
            products_data[product_key] = {
                "name": item["product_name"],
                "unit": item["product_unit"],
                "points": {}
            }
        
        point_id = item["point_id"]
        if point_id not in products_data[product_key]["points"]:
            products_data[product_key]["points"][point_id] = 0
        products_data[product_key]["points"][point_id] += item["count"]
    
    # Заполняем строки данных
    for product_id, data in sorted(products_data.items(), key=lambda x: x[1]["name"]):
        row = []
        product_name = f"{data['name']}"
        if data['unit']:
            product_name += f", {data['unit']}"
        row.append(product_name)
        
        for point_id, point_name, point_owner in points_list:
            count = data["points"].get(point_id, 0)
            row.append(count if count > 0 else "")
        
        ws.append(row)
    
    # Применяем стили к таблице
    apply_excel_styles(ws, points_list, products_data, supplier_items)
    
    # Сохраняем в BytesIO для отправки
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Собираем список ID элементов заказа для обновления статуса
    item_ids = [item["item_id"] for item in supplier_items]
    
    # Формируем имя файла
    excel_filename = f"Order_{supplier.name}_{product_type}_{moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Отправляем email через SMTP
    logger = logging.getLogger(__name__)
    
    try:
        # Для Mail.ru и многих других почтовых сервисов адрес отправителя должен совпадать
        # с адресом аутентификации. Если они не совпадают, используем SMTP_USER как отправителя.
        from_email = SMTP_FROM_EMAIL
        if SMTP_FROM_EMAIL.lower() != SMTP_USER.lower():
            logger.warning(f"SMTP_FROM_EMAIL ({SMTP_FROM_EMAIL}) не совпадает с SMTP_USER ({SMTP_USER}). "
                          f"Используем SMTP_USER как адрес отправителя для избежания ошибки 550.")
            from_email = SMTP_USER
        
        # Создаем сообщение
        msg = MIMEMultipart()
        
        # Правильно кодируем заголовок From
        # Mail.ru очень строгий к формату заголовков, поэтому используем простой формат
        # Если имя содержит только ASCII, добавляем его, иначе используем только email
        if SMTP_FROM_NAME:
            try:
                # Проверяем, содержит ли имя только ASCII символы
                SMTP_FROM_NAME.encode('ascii')
                # Если только ASCII, используем простой формат
                msg['From'] = f"{SMTP_FROM_NAME} <{from_email}>"
            except UnicodeEncodeError:
                # Mail.ru может отклонять заголовки с кириллицей, используем только email
                logger.warning(f"Имя отправителя содержит кириллицу, используем только email для избежания ошибки 550")
                msg['From'] = from_email
        else:
            msg['From'] = from_email
        
        msg['To'] = supplier.email
        
        # Добавляем адрес менеджера в Bcc, чтобы письмо попало в папку "Исходящие"
        # Это гарантирует, что копия письма будет сохранена в почтовом ящике менеджера
        msg['Bcc'] = from_email
        
        # Правильно кодируем заголовок Subject с поддержкой кириллицы
        subject_text = f"Заявка на товар - {product_type}"
        try:
            # Проверяем, содержит ли тема только ASCII символы
            subject_text.encode('ascii')
            msg['Subject'] = subject_text
        except UnicodeEncodeError:
            # Если содержит не-ASCII, кодируем через Header
            msg['Subject'] = Header(subject_text, 'utf-8')
        
        # Добавляем обязательные заголовки для совместимости с Mail.ru
        msg['Date'] = formatdate(localtime=True)
        msg['Message-ID'] = make_msgid()
        
        # Текст письма
        body = f"""Здравствуйте!
Во вложении файл с заявкой.
"""
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # Прикрепляем Excel файл
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(excel_buffer.read())
        encoders.encode_base64(part)
        
        # Кодируем имя файла для вложения
        # Mail.ru может быть строгим к формату заголовков, используем ASCII-safe имя
        ascii_filename = excel_filename.encode('ascii', 'ignore').decode('ascii') or 'application.xlsx'
        # Если ASCII имя получилось пустым или слишком коротким, используем дефолтное
        if not ascii_filename or len(ascii_filename) < 5:
            ascii_filename = f"application_{moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Используем простой формат для совместимости с Mail.ru
        part.add_header('Content-Disposition', f'attachment; filename="{ascii_filename}"')
        msg.attach(part)
        
        # Подключаемся к SMTP серверу и отправляем
        logger.info(f"Подключение к SMTP серверу: {SMTP_HOST}:{SMTP_PORT}")
        logger.info(f"Аутентификация как: {SMTP_USER}")
        logger.info(f"Отправка с адреса: {from_email}")
        logger.info(f"Отправка на адрес: {supplier.email}")
        
        if SMTP_PORT == 465:
            # SSL соединение
            server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT)
        else:
            # TLS соединение
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
            server.starttls()
        
        try:
            server.login(SMTP_USER, SMTP_PASSWORD)
            text = msg.as_string()
            
            # Формируем список получателей: основной адрес + Bcc (адрес менеджера для сохранения в "Исходящие")
            recipients = [supplier.email, from_email]
            
            # Используем from_email вместо SMTP_FROM_EMAIL для отправки
            # Bcc адрес включен в recipients, но не будет виден получателю (Bcc скрыт)
            refused = server.sendmail(from_email, recipients, text)
            if refused:
                logger.warning(f"Некоторые адреса были отклонены: {refused}")
            
            server.quit()
            logger.info(f"Email успешно отправлен на {supplier.email}")
        except smtplib.SMTPRecipientsRefused as e:
            logger.error(f"Ошибка: адреса получателей отклонены: {e}")
            server.quit()
            raise
        except smtplib.SMTPDataError as e:
            logger.error(f"Ошибка SMTP данных: {e}")
            logger.error(f"Код ошибки: {e.smtp_code}, сообщение: {e.smtp_error}")
            server.quit()
            raise
        except smtplib.SMTPException as e:
            logger.error(f"Общая SMTP ошибка: {e}")
            if hasattr(e, 'smtp_code'):
                logger.error(f"Код ошибки: {e.smtp_code}")
            if hasattr(e, 'smtp_error'):
                logger.error(f"Сообщение об ошибке: {e.smtp_error}")
            server.quit()
            raise
        
        # Помечаем товары как отправленные
        for item_id in item_ids:
            item = db.query(models.OrderItem).filter(models.OrderItem.id == item_id).first()
            if item:
                item.is_sent_to_supplier = True
                item.is_sending_to_supplier = False
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Заявка успешно отправлена на email {supplier.email}",
            "items_count": len(item_ids),
            "item_ids": item_ids,
            "status": "sent"
        }
        
    except smtplib.SMTPException as e:
        error_detail = str(e)
        if hasattr(e, 'smtp_code'):
            error_detail += f" (код: {e.smtp_code})"
        if hasattr(e, 'smtp_error'):
            error_detail += f" - {e.smtp_error}"
        logger.error(f"SMTP ошибка при отправке email: {error_detail}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при отправке email: {error_detail}")
    except Exception as e:
        logger.error(f"Ошибка при отправке email: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при отправке email: {str(e)}")

# ---- Orders ----
@app.get("/orders", response_model=list[schemas.Order])
@limiter.limit("5/second")
def read_orders(request: Request, status: str = None, db: Session = Depends(get_db)):
    query = db.query(models.Order).options(
        joinedload(models.Order.items).joinedload(models.OrderItem.product),
        joinedload(models.Order.point)
    )
    if status and status != "all":
        query = query.filter(models.Order.status == status)
    orders = query.order_by(models.Order.created_at.desc()).all()
    
    # Сериализуем заказы с unit из связанного product
    result = []
    for order in orders:
        order_dict = {
            "id": order.id,
            "point_id": order.point_id,
            "tg_name": order.tg_name,
            "tg_username": order.tg_username,
            "created_at": order.created_at,
            "status": order.status,
            "point": order.point,
            "items": [
                {
                    "id": item.id,
                    "type": item.type,
                    "subtype": item.subtype,
                    "name": item.name,
                    "count": item.count,
                    "product_id": item.product_id,
                    "suppliers_id": item.suppliers_id,
                    "unit": item.product.unit if item.product else None
                }
                for item in order.items
            ]
        }
        result.append(order_dict)
    return result

@app.get("/orders/pending-supplier", response_model=list[schemas.Order])
@limiter.limit("5/second")
def get_orders_pending_supplier(request: Request, db: Session = Depends(get_db)):
    """
    Получить заказы со статусом confirmed и sent_to_supplier = False,
    отображая только элементы заказа (order_items) с is_sent_to_supplier = False.
    
    У заказа может быть несколько поставщиков, и для части из них заявки отправлены,
    а для части еще нет. Показываем только те элементы, которые еще не отправлены.
    """
    # Загружаем заказы со статусом confirmed и sent_to_supplier = False
    orders = (
        db.query(models.Order)
        .options(
            joinedload(models.Order.items).joinedload(models.OrderItem.product),
            joinedload(models.Order.items).joinedload(models.OrderItem.supplier),
            joinedload(models.Order.point)
        )
        .filter(
            models.Order.status == "confirmed",
            models.Order.sent_to_supplier == False
        )
        .order_by(models.Order.created_at.desc())
        .all()
    )
    
    # Сериализуем заказы, фильтруя только элементы с is_sent_to_supplier = False
    result = []
    for order in orders:
        # Фильтруем элементы заказа: только те, которые еще не отправлены поставщику
        pending_items = [item for item in order.items if not item.is_sent_to_supplier]
        
        # Если после фильтрации нет элементов, пропускаем заказ
        if not pending_items:
            continue
        
        order_dict = {
            "id": order.id,
            "point_id": order.point_id,
            "tg_name": order.tg_name,
            "tg_username": order.tg_username,
            "created_at": order.created_at,
            "status": order.status,
            "point": order.point,
            "items": [
                {
                    "id": item.id,
                    "type": item.type,
                    "subtype": item.subtype,
                    "name": item.name,
                    "count": item.count,
                    "product_id": item.product_id,
                    "suppliers_id": item.suppliers_id,
                    "unit": item.product.unit if item.product else None
                }
                for item in pending_items
            ]
        }
        result.append(order_dict)
    
    return result

@app.get("/orders/{order_id}", response_model=schemas.Order)
@limiter.limit("5/second")
def get_order(request: Request, order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.Order).options(
        joinedload(models.Order.items).joinedload(models.OrderItem.product),
        joinedload(models.Order.point)
    ).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Сериализуем заказ с unit из связанного product
    order_dict = {
        "id": order.id,
        "point_id": order.point_id,
        "tg_name": order.tg_name,
        "tg_username": order.tg_username,
        "created_at": order.created_at,
        "status": order.status,
        "point": order.point,
        "items": [
            {
                "id": item.id,
                "type": item.type,
                "subtype": item.subtype,
                "name": item.name,
                "count": item.count,
                "product_id": item.product_id,
                "suppliers_id": item.suppliers_id,
                "unit": item.product.unit if item.product else None
            }
            for item in order.items
        ]
    }
    return order_dict

def is_duplicate_order(db: Session, order_data: schemas.OrderRequest) -> bool:
    """
    Проверка на дубликат заказа.
    
    Заказ считается дублем если:
    1. От той же точки (point_id)
    2. От того же пользователя (tg_username)
    3. Создан за последние 60 секунд
    4. Содержит идентичные товары (product_id + count)
    
    Args:
        db: Сессия базы данных
        order_data: Данные нового заказа для проверки
    
    Returns:
        True если заказ является дублем, False если уникальный
    """
    # Определяем временной порог (60 секунд назад)
    time_threshold = moscow_now() - timedelta(seconds=60)
    
    # Ищем недавние заказы от этой точки и пользователя
    recent_orders = db.query(models.Order).options(
        joinedload(models.Order.items)
    ).filter(
        models.Order.point_id == order_data.point_id,
        models.Order.tg_username == order_data.tg_username,
        models.Order.created_at > time_threshold
    ).all()
    
    if not recent_orders:
        return False
    
    # Формируем сигнатуру текущего заказа для сравнения
    # Сигнатура: множество строк вида "product_id:count"
    current_signature = set()
    for item in order_data.items:
        # Нормализуем product_id (может быть строкой или числом)
        product_id = item.product_id
        if isinstance(product_id, str):
            try:
                product_id = int(product_id) if product_id.strip() else None
            except (ValueError, TypeError):
                product_id = None
        
        # Создаем уникальный ключ: product_id + count
        key = f"{product_id}:{item.product_count}"
        current_signature.add(key)
    
    # Сравниваем с недавними заказами
    for order in recent_orders:
        order_signature = set()
        for item in order.items:
            key = f"{item.product_id}:{item.count}"
            order_signature.add(key)
        
        # Если сигнатуры совпадают - это дубль
        if current_signature == order_signature:
            logger.info(
                f"[ДУБЛЬ ОБНАРУЖЕН] Заказ от point_id={order_data.point_id}, "
                f"user={order_data.tg_username} является дублем заказа order_id={order.id}"
            )
            return True
    
    return False

@app.post("/orders")
@limiter.limit("5/second")
def create_order(
    request: Request,
    body: Any = Body(...),
    db: Session = Depends(get_db)
):
    """Создать заказы из массива объектов или одного объекта
    
    Принимает массив заказов или один заказ, где каждый заказ содержит:
    - point_id: ID точки продаж
    - tg_name: Имя клиента из Telegram
    - tg_username: Username клиента из Telegram
    - point: Название точки (не используется в БД)
    - items: Массив элементов заказа
    - text: Текст заказа (не используется в БД)
    
    Каждый элемент в items содержит:
    - product_id: ID товара (строка, конвертируется в int)
    - product_type: Тип товара
    - product_subtype: Подтип товара
    - product_name: Название товара
    - product_count: Количество товара
    - product_unit: Единица измерения товара (опционально)
    - suppliers_id: ID поставщика (строка, конвертируется в int)
    - suppliers_name: Название поставщика (не используется в БД)
    """
    try:
        # Нормализуем входные данные: если пришел один объект, оборачиваем в массив
        if isinstance(body, dict):
            # Если это словарь (один объект), оборачиваем в массив
            try:
                # Пытаемся валидировать как один объект
                order_request = schemas.OrderRequest(**body)
                orders = [order_request]
            except ValidationError as e:
                logger.error(f"[422] Ошибка валидации заказа (dict). Body: {body}. Ошибки: {e.errors()}")
                raise HTTPException(status_code=422, detail=e.errors())
            except Exception as e:
                logger.error(f"[422] Ошибка парсинга заказа (dict). Body: {body}. Ошибка: {str(e)}")
                raise HTTPException(status_code=422, detail=f"Invalid order data: {str(e)}")
        elif isinstance(body, list):
            # Если это массив, валидируем каждый элемент
            orders = []
            for item in body:
                if isinstance(item, dict):
                    try:
                        orders.append(schemas.OrderRequest(**item))
                    except ValidationError as e:
                        logger.error(f"[422] Ошибка валидации заказа (list item). Item: {item}. Ошибки: {e.errors()}")
                        raise HTTPException(status_code=422, detail=e.errors())
                elif isinstance(item, schemas.OrderRequest):
                    orders.append(item)
                else:
                    raise HTTPException(status_code=422, detail="Array items must be objects")
        else:
            raise HTTPException(status_code=422, detail="Body must be either an object or an array of objects")
        
        created_orders = []
        
        for order_data in orders:
            # 🛡️ Проверка на дубликат заказа
            if is_duplicate_order(db, order_data):
                logger.warning(
                    f"[ДУБЛЬ ПРОПУЩЕН] Пропущен дубликат заказа от "
                    f"point_id={order_data.point_id}, user={order_data.tg_username}"
                )
                # Возвращаем успешный ответ, но заказ НЕ создаем
                continue
            
            # Создаем заказ
            db_order = models.Order(
                point_id=order_data.point_id,
                tg_name=order_data.tg_name,
                tg_username=order_data.tg_username
            )
            db.add(db_order)
            db.flush()  # Получаем ID заказа без commit
            
            # Создаем элементы заказа
            for item_data in order_data.items:
                # Конвертируем ID в int (обрабатываем строки, числа, пустые строки и None)
                product_id = None
                if item_data.product_id is not None:
                    if isinstance(item_data.product_id, int):
                        product_id = item_data.product_id
                    elif isinstance(item_data.product_id, str) and item_data.product_id.strip():
                        try:
                            product_id = int(item_data.product_id)
                        except (ValueError, TypeError):
                            product_id = None
                
                suppliers_id = None
                if item_data.suppliers_id is not None:
                    if isinstance(item_data.suppliers_id, int):
                        suppliers_id = item_data.suppliers_id
                    elif isinstance(item_data.suppliers_id, str) and item_data.suppliers_id.strip():
                        try:
                            suppliers_id = int(item_data.suppliers_id)
                        except (ValueError, TypeError):
                            suppliers_id = None
                
                db_item = models.OrderItem(
                    order_id=db_order.id,
                    product_id=product_id,
                    suppliers_id=suppliers_id,
                    type=item_data.product_type,
                    subtype=item_data.product_subtype,
                    name=item_data.product_name,
                    count=item_data.product_count
                )
                db.add(db_item)
            
            created_orders.append(db_order)
        
        db.commit()
        
        # Обновляем объекты для возврата и загружаем связанные данные
        result_orders = []
        for db_order in created_orders:
            # Перезагружаем заказ с связанными данными
            order_with_relations = (
                db.query(models.Order)
                .options(
                    joinedload(models.Order.items).joinedload(models.OrderItem.product),
                    joinedload(models.Order.point)
                )
                .filter(models.Order.id == db_order.id)
                .first()
            )
            if order_with_relations:
                # Сериализуем заказ в словарь (как в других эндпоинтах)
                order_dict = {
                    "id": order_with_relations.id,
                    "point_id": order_with_relations.point_id,
                    "tg_name": order_with_relations.tg_name,
                    "tg_username": order_with_relations.tg_username,
                    "created_at": order_with_relations.created_at,
                    "status": order_with_relations.status,
                    "point": order_with_relations.point,
                    "manager_telegram_id": order_with_relations.point.manager_telegram_id if order_with_relations.point else None,
                    "items": [
                        {
                            "id": item.id,
                            "type": item.type,
                            "subtype": item.subtype,
                            "name": item.name,
                            "count": item.count,
                            "product_id": item.product_id,
                            "suppliers_id": item.suppliers_id,
                            "unit": item.product.unit if item.product else None
                        }
                        for item in order_with_relations.items
                    ]
                }
                result_orders.append(order_dict)
        
        return result_orders
    except HTTPException:
        # Пробрасываем HTTP исключения как есть
        raise
    except Exception as e:
        # Логируем все остальные ошибки
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error creating order: {str(e)}")
        print(f"Traceback: {error_trace}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.patch("/orders/{order_id}/status")
def update_order_status(order_id: int, status: str = Query(...), db: Session = Depends(get_db)):
    """Обновить статус заказа"""
    if status not in ["new", "confirmed", "rejected"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    order.status = status
    db.commit()
    db.refresh(order)
    return order

# ---- Export to Excel ----
@app.get("/orders/{order_id}/export")
def export_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Order_{order.id}"

    ws["A1"] = f"Заказ №{order.id}"
    ws.append([""])
    ws.append(["Точка:", order.point.name if order.point else "—"])
    ws.append(["Дата:", order.created_at.strftime("%d.%m.%Y, %H:%M:%S")])
    ws.append(["Имя TG:", order.tg_name or "—"])
    ws.append(["Username TG:", order.tg_username or "—"])
    ws.append(["Статус:", order.status])
    ws.append([""])
    ws.append(["Продукты:"])
    ws.append(["Тип", "Подтип", "Название", "Кол-во"])

    for item in order.items:
        ws.append([item.type, item.subtype, item.name, item.count])

    filename = f"order_{order.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("exports", filename)
    os.makedirs("exports", exist_ok=True)
    wb.save(filepath)

    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/export/pack")
def export_pack():
    """
    Генерирует и заполняет шаблон 'Упаковка' на основе подтверждённых заказов из БД.
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только упаковку
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "упаковка"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон перед заполнением
        subtypes = {
            "05 Посуда одноразовая": {
                "unit": "шт.",
                "items": [
                    'Вилка "Кристалл"',
                    'Ложка "Кристалл"',
                    'Ложка чайная',
                    'Нож "Кристалл"',
                    'Палочка для размешивания',
                    'Стакан 200 мл. прозрачный'
                ]
            },
            "07 Хозтовары": {
                "unit": "рул.",
                "items": [
                    "Шпагат ПП 2200 текс (баб по 4,5-5кг.)"
                ]
            },
            "Заказные": {
                "unit": "шт.",
                "items": [
                    "Пакет майка 'Добрые булки'",
                    "Стикеры мал.",
                    "Стикеры бол."
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="Упаковка")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

@app.get("/export/limonad")
def export_limonad():
    """
    Генерирует и заполняет шаблон 'Лимонад' (горизонтальный формат).
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только лимонад
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "лимонад"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон (только unsub)
        subtypes = {
            "unsub": {
                "unit": "шт.",
                "items": [
                    "Фрустайл лимон-лайм ж/б",
                    "Фрустайл апельсин ж/б",
                    "Эвервес ж/б",
                    "Эвервес 0,5",
                    "Эвервес б/сах 0,5",
                    "Липтон зеленый",
                    "Липтон лимон черный",
                    "Вода газ",
                    "любимый/ябл",
                    "любимый/вишн",
                    "любимый/клуб",
                    "Вода газ 2л",
                    "Вода б/газа 2л"
                ]
            }
        }

        generate_template_horizontal(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template_horizontal(orders, packaging_type="Лимонад")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

@app.get("/export/food")
def export_food():
    """
    Генерирует и заполняет шаблон 'Часть продуктов' на основе подтверждённых заказов.
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только часть продуктов
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "часть продуктов"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон
        subtypes = {
            "unsub": {
                "unit": "шт.",
                "items": [
                    "масло растительное",
                    "соль",
                    "яйцо",
                    "чеснок",
                    "сок яблоко",
                    "сок персик/абрик",
                    "сок мульти",
                    "сок вишня",
                    "сок апельсин",
                    "чай черн",
                    "чай зелен"
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="часть продуктов")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

# ---- NEW ----
@app.get("/export/foodSupply")
def export_food():
    """
    Генерирует и заполняет шаблон 'Часть продуктов' на основе подтверждённых заказов.
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только часть продуктов
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "пищеснаб"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон
        subtypes = {
            "unsub": {
                "unit": "шт.",
                "items": [
                    "конфитюр вишня",
                    "конфитюр яблоко",
                    "конфитюр малина",
                    "конфитюр смородина",
                    "изюм",
                    "мак",
                    "маргарин"
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="пищеснаб")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

@app.get("/export/milk")
def export_food():
    """
    Генерирует и заполняет шаблон 'Часть продуктов' на основе подтверждённых заказов.
    """
    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только часть продуктов
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "молочка"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон
        subtypes = {
            "unsub": {
                "unit": "шт.",
                "items": [
                    "творог",
                    "сметана"
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="молочка")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

@app.get("/export/sausages")
def export_food():
    """
    Генерирует и заполняет шаблон 'Часть продуктов' на основе подтверждённых заказов.
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только часть продуктов
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "сосиски"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон
        subtypes = {
            "unsub": {
                "unit": "шт.",
                "items": [
                    "сосиски"
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="сосиски")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

@app.get("/export/water")
def export_food():
    """
    Генерирует и заполняет шаблон 'Часть продуктов' на основе подтверждённых заказов.
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только часть продуктов
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "вода"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон
        subtypes = {
            "unsub": {
                "unit": "шт.",
                "items": [
                    "Вода 19л"
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="вода")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

@app.get("/export/flour")
def export_food():
    """
    Генерирует и заполняет шаблон 'Часть продуктов' на основе подтверждённых заказов.
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только часть продуктов
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "мука/сахар"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон
        subtypes = {
            "unsub": {
                "unit": "шт.",
                "items": [
                    "мука",
                    "сахар"
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="мука/сахар")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

@app.get("/export/pone")
def export_food():
    """
    Генерирует и заполняет шаблон 'Часть продуктов' на основе подтверждённых заказов.
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только часть продуктов
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "ирексол софт сдоба"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон
        subtypes = {
            "unsub": {
                "unit": "шт.",
                "items": [
                    "ирексол софт сдоба"
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="ирексол софт сдоба")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

@app.get("/export/whiteness")
def export_food():
    """
    Генерирует и заполняет шаблон 'Часть продуктов' на основе подтверждённых заказов.
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только часть продуктов
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "белизна"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон
        subtypes = {
            "unsub": {
                "unit": "товар",
                "items": [
                    "Белизна",
                    "Губка для посуд 10 шт/уп",
                    "ЖМС универсал 5 л",
                    "Крот 1  л",
                    "МОП микрофибра шт",
                    "Ника 5 л",
                    "Освежитель воздуха",
                    "Пемолюкс",
                    "Ср-во для гриля ",
                    "Ср-во для стекол",
                    "туалетная бумага/рулоны",
                    "антисептик",
                    "жидкое мыло",
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="белизна")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()

@app.get("/export/waffleFabric")
def export_food():
    """
    Генерирует и заполняет шаблон 'Часть продуктов' на основе подтверждённых заказов.
    """

    # 1️⃣ — Загружаем заказы
    db = SessionLocal()
    try:
        orders = (
            db.query(models.Order)
            .options(
                joinedload(models.Order.point),
                joinedload(models.Order.items)
            )
            .filter(models.Order.status == "confirmed", models.Order.sent_to_supplier == False)
            .all()
        )

        # 2️⃣ — Фильтруем только часть продуктов
        for order in orders:
            order.items = [i for i in order.items if i.type and i.type.lower() == "вафел.полотно"]

        if not orders:
            return {"message": "Нет данных для заполнения шаблона"}

        # 3️⃣ — Создаём шаблон
        subtypes = {
            "unsub": {
                "unit": "шт.",
                "items": [
                    "Вафел.полотно",
                    "Мешки д/мусора 120 л шт",
                    "Мешки д/мусора 240 л шт",
                    "Мешки д/мусора 60 л шт",
                    "Пакет фас 18*27 уп",
                    "Пакет фас 22*32 уп",
                    "Пакет фас 25*40 уп",
                    "Пергамент (40*25)",
                    "Перчатки винил 100шт/уп",
                    "Перчатки полиэт.100шт/уп",
                    "Перчатки хозяйственные",
                    "Салфетки 25*25 49 уп/коробка",
                    "Фартук однораз упак",
                    "Шапочка 100шт/уп"
                ]
            }
        }

        generate_template(subtypes)

        # 4️⃣ — Заполняем шаблон
        filled_path = fill_template(orders, packaging_type="Вафел.полотно")

        order_item_ids = [order.id for order in orders]

        # 5️⃣ — Создаём запись в истории (со статусом "отказана")
        history_record = models.History(
            file_path=filled_path,
            created_at=moscow_now()
        )
        db.add(history_record)
        db.commit()  # обязательно commit, чтобы id был присвоен

        # 5️⃣ — Возвращаем файл
        filename = f"foodSupply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = FileResponse(
            filled_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Order-Item-IDs"] = ",".join(map(str, order_item_ids))
        response.headers["X-History-ID"] = str(history_record.id)  # <-- добавили id истории
        return response
    finally:
        db.close()
